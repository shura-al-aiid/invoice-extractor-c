"""Serialise an :class:`ExtractionResult` to Excel or CSV.

The Excel workbook has three sheets:

* **Extracted Invoices** — the clean table, one row per invoice. Numbers are real
  numbers and dates are ISO strings, so the file is analysis-ready.
* **Flagged for Review** — every item a human needs to resolve, with the reason.
* **Summary** — processing counts (invoices, fields filled, flags).

CSV output is intentionally just the clean table (the format has no concept of
multiple sheets); the flags are written to a sibling ``*_flagged.csv``.
"""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .models import INVOICE_COLUMNS, ExtractionResult, Flag, Invoice

# Shared styling constants.
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_FLAG_HEADER_FILL = PatternFill("solid", fgColor="C0392B")
_MONEY_FORMAT = "#,##0.00"
# 1-based indices of the money columns, derived from the canonical column order so
# this stays correct if columns are added/reordered in models.INVOICE_COLUMNS.
_MONEY_COLUMNS = tuple(
    i for i, name in enumerate(INVOICE_COLUMNS, start=1)
    if name in ("subtotal", "tax", "total")
)


# --------------------------------------------------------------------------- #
# Public entry points
# --------------------------------------------------------------------------- #
def write_xlsx(result: ExtractionResult, path: Path) -> None:
    """Write the full three-sheet workbook to ``path``."""
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    invoices_ws = wb.active
    invoices_ws.title = "Extracted Invoices"
    _write_invoices_sheet(invoices_ws, result)

    flags_ws = wb.create_sheet("Flagged for Review")
    _write_flags_sheet(flags_ws, result)

    summary_ws = wb.create_sheet("Summary")
    _write_summary_sheet(summary_ws, result)

    wb.save(path)


def write_csv(result: ExtractionResult, path: Path) -> Path:
    """Write the clean table to ``path`` and flags to ``*_flagged.csv``.

    Returns the path of the flags file so the CLI can report it.
    """
    path.parent.mkdir(parents=True, exist_ok=True)

    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(Invoice.headers())
        for invoice in result.invoices:
            writer.writerow(_blank_none(invoice.as_row()))

    flagged_path = path.with_name(path.stem + "_flagged.csv")
    with flagged_path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(Flag.headers())
        for flag in result.flags:
            writer.writerow(flag.as_row())

    return flagged_path


# --------------------------------------------------------------------------- #
# Sheet builders
# --------------------------------------------------------------------------- #
def _write_invoices_sheet(ws: Worksheet, result: ExtractionResult) -> None:
    _write_header(ws, Invoice.headers(), _HEADER_FILL)
    for invoice in result.invoices:
        ws.append(_blank_none(invoice.as_row()))

    # Apply a real number format to the money columns so values stay numeric.
    for row in range(2, ws.max_row + 1):
        for col in _MONEY_COLUMNS:
            ws.cell(row=row, column=col).number_format = _MONEY_FORMAT

    _finalise(ws, len(Invoice.headers()))


def _write_flags_sheet(ws: Worksheet, result: ExtractionResult) -> None:
    _write_header(ws, Flag.headers(), _FLAG_HEADER_FILL)
    if result.flags:
        for flag in result.flags:
            ws.append(flag.as_row())
    else:
        ws.append(["— none —", "", "", ""])
    _finalise(ws, len(Flag.headers()))


def _write_summary_sheet(ws: Worksheet, result: ExtractionResult) -> None:
    _write_header(ws, ["Metric", "Value"], _HEADER_FILL)

    total_possible = result.invoice_count * result.extractable_field_count
    total_filled = sum(result.fields_filled(inv) for inv in result.invoices)
    avg_fields = (total_filled / result.invoice_count) if result.invoice_count else 0
    flagged_invoices = {flag.source_file for flag in result.flags}

    rows = [
        ("Invoices processed", result.invoice_count),
        ("Extractable fields per invoice", result.extractable_field_count),
        ("Fields filled (all invoices)", f"{total_filled} / {total_possible}"),
        ("Average fields filled per invoice", f"{avg_fields:.1f}"),
        ("Invoices with at least one flag", len(flagged_invoices)),
        ("Total review flags raised", result.flag_count),
    ]
    for metric, value in rows:
        ws.append([metric, value])

    # Per-invoice fields-filled breakdown for transparency.
    ws.append([])
    ws.append(["Per-invoice: fields filled", ""])
    breakdown_header_row = ws.max_row
    ws.cell(row=breakdown_header_row, column=1).font = Font(bold=True)
    for invoice in result.invoices:
        ws.append(
            [
                invoice.source_file,
                f"{result.fields_filled(invoice)} / {result.extractable_field_count}",
            ]
        )

    _finalise(ws, 2)


# --------------------------------------------------------------------------- #
# Styling helpers
# --------------------------------------------------------------------------- #
def _write_header(ws: Worksheet, headers: list[str], fill: PatternFill) -> None:
    ws.append(headers)
    for col, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")


def _finalise(ws: Worksheet, num_columns: int) -> None:
    """Freeze the header row and auto-size columns to their content."""
    ws.freeze_panes = "A2"
    for col in range(1, num_columns + 1):
        letter = get_column_letter(col)
        widest = 0
        for cell in ws[letter]:
            if cell.value is not None:
                widest = max(widest, len(str(cell.value)))
        # Clamp width to a sensible range so long text doesn't explode the layout.
        ws.column_dimensions[letter].width = min(max(widest + 2, 12), 50)


def _blank_none(row: list) -> list:
    """Render ``None`` as an empty cell (the visible signal of a withheld value)."""
    return ["" if value is None else value for value in row]
